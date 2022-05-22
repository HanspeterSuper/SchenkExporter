from openpyxl import load_workbook
from zoneinfo import ZoneInfo
import datetime
from dotenv import load_dotenv
import os
from db import create_connection
from db import execute_read_query
from db import execute_query

def makeexcel(user_email, kw, jahr):

    currentDirectory = os.getcwd()

    currentDirectory = currentDirectory + '/tmp'

    for f in os.listdir(currentDirectory):
        os.remove(os.path.join(currentDirectory, f))

    load_dotenv()

    email = user_email.replace(".", " ")

    connectiondb = create_connection(
        os.getenv('MYSQL_HOST'), 
        os.getenv('MYSQL_USER'), 
        os.getenv('MYSQL_PASSWORD'), 
        os.getenv('MYSQL_DB')
    )

    select_user = "SELECT id, alias, username FROM kimai2_users WHERE alias='{}'".format(email)
    user = execute_read_query(connectiondb, select_user)
    user = user[0]
    
    user_alias = user[1]
    
    user_id = user[0]

    select_ferienGuthaben = "SELECT value FROM kimai2_user_preferences WHERE user_id='{}' AND name='ferien_guthaben'".format(user_id)
    ferienGuthaben = execute_read_query(connectiondb, select_ferienGuthaben)
    ferienGuthaben = ferienGuthaben[0]
    ferienGuthaben = ferienGuthaben[0]

    kw_jahr = [jahr, kw]

    if int(kw) - 1 < 0:
        select_vorwoche_db = "SELECT gesamtTotal FROM schenkExporter WHERE user='{}' AND kw_jahr='{}'".format(user_id, str(int(jahr)-1) + "_" + str(52))
    else:
        select_vorwoche_db = "SELECT gesamtTotal FROM schenkExporter WHERE user='{}' AND kw_jahr='{}'".format(user_id, jahr + "_" + str(int(kw)-1))
    vorwoche_db = execute_read_query(connectiondb, select_vorwoche_db)
    try:
        vorwoche_db = vorwoche_db[0]
    except IndexError:
        vorwoche_db.insert(0, '_')  
    vorwoche_db = vorwoche_db[0]        
    if vorwoche_db != '_':
        vorwoche = vorwoche_db
    else:
        select_vorwoche = "SELECT value FROM kimai2_user_preferences WHERE user_id='{}' AND name='total_vorwoche'".format(user_id)
        vorwoche = execute_read_query(connectiondb, select_vorwoche)
        vorwoche = vorwoche[0]
        vorwoche = vorwoche[0]

    if int(kw) - 1 < 0:
        select_ferienGuthaben_db = "SELECT ferienGuthaben FROM schenkExporter WHERE user='{}' AND kw_jahr='{}'".format(user_id, str(int(jahr)-1) + "_" + str(52))
        newYear = 20
    else:
        select_ferienGuthaben_db = "SELECT ferienGuthaben FROM schenkExporter WHERE user='{}' AND kw_jahr='{}'".format(user_id, jahr + "_" + str(int(kw)-1))
        newYear = 0
    ferienGuthaben_db = execute_read_query(connectiondb, select_ferienGuthaben_db)
    try:
        ferienGuthaben_db = ferienGuthaben_db[0]
    except IndexError:
        ferienGuthaben_db.insert(0, '_')  
    ferienGuthaben_db = ferienGuthaben_db[0]        
    if ferienGuthaben_db != '_':
        ferienGuthaben = str(int(ferienGuthaben_db) + newYear)
    else:
        select_ferienGuthaben = "SELECT value FROM kimai2_user_preferences WHERE user_id='{}' AND name='ferien_guthaben'".format(user_id)
        ferienGuthaben = execute_read_query(connectiondb, select_ferienGuthaben)
        ferienGuthaben = ferienGuthaben[0]
        ferienGuthaben = str(int(ferienGuthaben[0]) + newYear)

    wb = load_workbook(filename = './VORLAGE_Kimai.xlsx')
    zeitrapport_ranges = wb['Zeitrapport']
    arbeitsRapporte_ranges = wb['Arbeits Rapporte']
    inendienst_ranges = wb['Innendienst']
    zeitrapport_ranges['B2'].value = user_alias
    zeitrapport_ranges['J2'].value = kw_jahr[1]
    zeitrapport_ranges['P14'].value = ferienGuthaben
    zeitrapport_ranges['K15'].value = vorwoche

    DateDay = []

    for i in range(7):
        d = datetime.datetime.strptime(str(kw_jahr[0]) + '-W' + str(kw_jahr[1]) + '-' + str(i), "%Y-W%W-%w")
        DateDay.insert(i, d.strftime('%d.%m.%Y'))
    DateDay.sort()

    zeitrapport_ranges['M2'].value = datetime.datetime.strptime(DateDay[0],'%d.%m.%Y')
    zeitrapport_ranges['O2'].value = datetime.datetime.strptime(DateDay[6],'%d.%m.%Y')

    for i in range(7):
        zeitrapport_ranges['B' + str(i + 5)].value = datetime.datetime.strptime(DateDay[i],'%d.%m.%Y')

    select_fahrzeit_projId = "SELECT id FROM kimai2_projects WHERE name='Fahrzeit'"
    fahrzeit_projId = execute_read_query(connectiondb, select_fahrzeit_projId)
    fahrzeit_projId = fahrzeit_projId[0][0]

    select_publicHoliday_projId = "SELECT id FROM kimai2_projects WHERE name='Public Holiday'"
    publicHoliday_projId = execute_read_query(connectiondb, select_publicHoliday_projId)
    publicHoliday_projId = publicHoliday_projId[0][0]

    select_vacation_projId = "SELECT id FROM kimai2_projects WHERE name='Vacation'"
    vacation_projId = execute_read_query(connectiondb, select_vacation_projId)
    vacation_projId = vacation_projId[0][0]

    select_sick_projId = "SELECT id FROM kimai2_projects WHERE name='Sick'"
    sick_projId = execute_read_query(connectiondb, select_sick_projId)
    sick_projId = sick_projId[0][0]

    first_row_of_each_day = []

    for i in range(7):
        select_first_row_of_each_day = "SELECT id, start_time, project_id FROM kimai2_timesheet WHERE user='{}' AND start_time<'{}' AND start_time>'{}' LIMIT 1".format(user_id, datetime.datetime.strptime(DateDay[i] + ' 23:59:59', '%d.%m.%Y %H:%M:%S'), datetime.datetime.strptime(DateDay[i], '%d.%m.%Y'))
        first_row_of_each_day.insert(i, execute_read_query(connectiondb, select_first_row_of_each_day))
        try:
            first_row_of_each_day[i] = first_row_of_each_day[i][0]
        except IndexError:
            first_row_of_each_day[i].insert(0, '_')
            first_row_of_each_day[i] = first_row_of_each_day[i][0]

    last_row_of_each_day = []

    for i in range(7):
        select_last_row_of_each_day = "SELECT id, end_time, project_id FROM kimai2_timesheet WHERE user='{}' AND end_time<'{}' AND end_time>'{}' ORDER BY start_time DESC LIMIT 1".format(user_id, datetime.datetime.strptime(DateDay[i] + ' 23:59:59', '%d.%m.%Y %H:%M:%S'), datetime.datetime.strptime(DateDay[i], '%d.%m.%Y'))
        last_row_of_each_day.insert(i, execute_read_query(connectiondb, select_last_row_of_each_day))
        try:
            last_row_of_each_day[i] = last_row_of_each_day[i][0]
        except IndexError:
            last_row_of_each_day[i].insert(0, '_')
            last_row_of_each_day[i] = last_row_of_each_day[i][0]

    position_of_innendienst = 6
    total_work_of_the_week = 0
    total_vacation_of_the_week = 0

    for i in range(7):
        if first_row_of_each_day[i] != "_":
            if (first_row_of_each_day[i][2] != publicHoliday_projId) and (first_row_of_each_day[i][2] != vacation_projId) and (first_row_of_each_day[i][2] != sick_projId):
                timesheet_of_the_day = []
                zeitrapport_ranges['C' + str(i + 5)].value = first_row_of_each_day[i][1].replace(tzinfo=ZoneInfo('UTC')).astimezone(ZoneInfo('Europe/Zurich')).time().replace(tzinfo=None)
                zeitrapport_ranges['D' + str(i + 5)].value = last_row_of_each_day[i][1].replace(tzinfo=ZoneInfo('UTC')).astimezone(ZoneInfo('Europe/Zurich')).time().replace(tzinfo=None)
                select_timesheet_of_the_day = "SELECT kimai2_timesheet.duration, kimai2_timesheet.project_id, kimai2_timesheet_meta.value, kimai2_timesheet.start_time FROM kimai2_timesheet LEFT JOIN kimai2_timesheet_meta ON kimai2_timesheet_meta.timesheet_id=kimai2_timesheet.id WHERE user='{}' AND kimai2_timesheet.end_time<='{}' AND kimai2_timesheet.start_time>='{}' AND kimai2_timesheet_meta.name='ticket_number' ORDER BY kimai2_timesheet.start_time DESC ".format(user_id, last_row_of_each_day[i][1].strftime('%Y-%m-%d %H:%M:%S'), first_row_of_each_day[i][1].strftime('%Y-%m-%d %H:%M:%S'))
                timesheet_of_the_day.insert(i, execute_read_query(connectiondb, select_timesheet_of_the_day))
                timesheet_of_the_day = timesheet_of_the_day[0]

                total_work_on_day = 0
                for j in timesheet_of_the_day:
                    total_work_on_day = total_work_on_day + j[0]
                total_work_on_day = total_work_on_day/3600
                work_on_day_without_break = last_row_of_each_day[i][1] - first_row_of_each_day[i][1]
                hours,minutes,seconds = str(work_on_day_without_break).split(':')
                work_on_day_without_break = (int(hours) * 3600 + int(minutes) * 60 + int(seconds))/3600
                zeitrapport_ranges['E' + str(i + 5)].value = work_on_day_without_break - total_work_on_day

                total_work_of_the_week = total_work_of_the_week + total_work_on_day

                total_id_of_day = 0
                index_j = 0
                actual_ticket_number = 999999
                position_of_ticket = 6
                for j in timesheet_of_the_day:
                    if j[1] == fahrzeit_projId:
                        pass
                    elif j[2] == None:
                        total_id_of_day = total_id_of_day + j[0]
                    else:
                        if actual_ticket_number == j[2]:
                            pass
                        else:
                            actual_ticket_number = j[2]
                            total_time_for_ticket = j[0]
                            index_h = 0
                            for h in timesheet_of_the_day:
                                if index_h == index_j:
                                    pass
                                elif actual_ticket_number == h[2]:
                                    total_time_for_ticket = total_time_for_ticket + h[0]
                                index_h = index_h + 1
                            arbeitsRapporte_ranges.cell(row=position_of_ticket, column=i*2+1).value = actual_ticket_number
                            arbeitsRapporte_ranges.cell(row=position_of_ticket, column=i*2+2).value = total_time_for_ticket/3600
                            position_of_ticket = position_of_ticket + 1
                    index_j = index_j + 1

                if total_id_of_day > 0:
                    inendienst_ranges['A' + str(position_of_innendienst)].value = j[3].date()
                    inendienst_ranges['D' + str(position_of_innendienst)].value = "Gemäss Kimai2"
                    inendienst_ranges['Q' + str(position_of_innendienst)].value = total_id_of_day/3600
                    zeitrapport_ranges['H' + str(i + 5)].value = total_id_of_day/3600
                    position_of_innendienst = position_of_innendienst + 1

            elif first_row_of_each_day[i][2] == publicHoliday_projId:
                total_work_of_the_week = total_work_of_the_week + 8.5
                arbeitsRapporte_ranges['I' + str(i + 28)].value = "X"
            elif first_row_of_each_day[i][2] == vacation_projId:
                total_work_of_the_week = total_work_of_the_week + 8.5
                arbeitsRapporte_ranges['G' + str(i + 28)].value = "X"
                total_vacation_of_the_week = total_vacation_of_the_week + 1
            elif first_row_of_each_day[i][2] == sick_projId:
                total_work_of_the_week = total_work_of_the_week + 8.5
                arbeitsRapporte_ranges['H' + str(i + 28)].value = "X"

    gesamtTotal = float(vorwoche) + (total_work_of_the_week - 42.5)
    zeitrapport_ranges["K16"].value = gesamtTotal

    update_vorwoche = "UPDATE kimai2_user_preferences SET value = '{}' WHERE user_id='{}' AND name='total_vorwoche'".format(gesamtTotal, user_id)
    execute_query(connectiondb, update_vorwoche)

    newFerienGuthaben = int(ferienGuthaben) - total_vacation_of_the_week
    zeitrapport_ranges["P16"].value = newFerienGuthaben

    update_ferienGuthaben = "UPDATE kimai2_user_preferences SET value = '{}' WHERE user_id='{}' AND name='ferien_guthaben'".format(newFerienGuthaben, user_id)
    execute_query(connectiondb, update_ferienGuthaben)
    
    update_gesamtTotal_db = "INSERT INTO schenkExporter SET gesamtTotal = '{}', user='{}', kw_jahr = '{}', ferienGuthaben = '{}' ON DUPLICATE KEY UPDATE gesamtTotal = '{}', user='{}', kw_jahr = '{}', ferienGuthaben = '{}'".format(gesamtTotal, user_id, str(jahr) + "_" + str(kw), newFerienGuthaben, gesamtTotal, user_id, str(jahr) + "_" + str(kw), newFerienGuthaben)
    execute_query(connectiondb, update_gesamtTotal_db)       
    
    excelFilename = 'tmp/{}_KW{}_{}.xlsx'.format(user[2], kw_jahr[1], kw_jahr[0])

    wb.save(filename = excelFilename)

    return excelFilename